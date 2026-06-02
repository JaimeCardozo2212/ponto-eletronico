"""
Página 6 — Exportar Excel
Filters, preview, and formatted XLSX download with 3 sheets.
"""

import streamlit as st
from datetime import date, datetime
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

from database import (
    get_time_entries,
    get_projects,
    get_allocations_summary,
    get_company_allocations_summary,
    get_allocations_for_entry,
)
from utils import (
    parse_time,
    calculate_worked_hours,
    format_hours,
    get_daily_hours_with_projects,
    get_weekday_name,
)

st.title("📥 Exportar para Excel")
st.caption("Filtre os dados e baixe um arquivo Excel formatado")

# ── Filters ──
col_f1, col_f2 = st.columns(2)
with col_f1:
    date_range = st.date_input(
        "📅 Período",
        value=(date.today().replace(day=1), date.today()),
        format="DD/MM/YYYY",
        key="exp_date",
    )
with col_f2:
    agg_option = st.selectbox(
        "📊 Agrupamento",
        ["Diário (detalhado)", "Semanal", "Mensal"],
        key="exp_agg",
    )

# Parse date range
if isinstance(date_range, tuple) and len(date_range) == 2:
    start_date, end_date = date_range
elif isinstance(date_range, date):
    start_date = end_date = date_range
else:
    start_date = date.today().replace(day=1)
    end_date = date.today()

# Fetch data
entries = get_time_entries(start_date, end_date, order_desc=False)
enriched = [get_daily_hours_with_projects(e) for e in entries]

if not enriched:
    st.warning("Nenhum registro encontrado no período selecionado.")
    st.stop()

# ── Build preview DataFrame ──
preview_data = []
for e in enriched:
    d = date.fromisoformat(e["date"]) if isinstance(e["date"], str) else e["date"]
    preview_data.append({
        "Data": d.strftime("%d/%m/%Y"),
        "Dia": get_weekday_name(d),
        "Entrada": e["start_time"] or "—",
        "Início Almoço": e["lunch_start"] or "—",
        "Volta Almoço": e["lunch_end"] or "—",
        "Saída": e["end_time"] or "—",
        "Horas": e["worked_hours"],
        "Projetos": e["project_names"],
        "Empresas": e["company_names"],
        "Observações": (e["notes"] or "") if e["notes"] else "",
    })

df_preview = pd.DataFrame(preview_data)

st.subheader("👁️ Preview")
st.dataframe(df_preview, use_container_width=True, hide_index=True)

total_hours = sum(e["worked_hours"] for e in enriched)
st.caption(f"**Total:** {format_hours(total_hours)} | **Registros:** {len(enriched)}")

# ── Excel Generation ──
st.markdown("---")

def generate_excel() -> BytesIO:
    """Generate a formatted Excel workbook with 3 sheets."""
    wb = Workbook()

    # Styles
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    cell_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    bold_font = Font(bold=True)

    def style_header(ws, headers, row=1):
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

    def style_data_cell(ws, row, col, value, fmt=None):
        cell = ws.cell(row=row, column=col, value=value)
        cell.alignment = cell_alignment
        cell.border = thin_border
        return cell

    # ── Sheet 1: Registros Detalhados ──
    ws1 = wb.active
    ws1.title = "Registros Detalhados"
    headers1 = [
        "Data", "Dia", "Entrada", "Início Almoço", "Volta Almoço",
        "Saída", "Horas Trabalhadas", "Projetos", "Empresas", "Observações",
    ]
    style_header(ws1, headers1)

    for i, row_data in enumerate(preview_data, 2):
        style_data_cell(ws1, i, 1, row_data["Data"])
        style_data_cell(ws1, i, 2, row_data["Dia"])
        style_data_cell(ws1, i, 3, row_data["Entrada"])
        style_data_cell(ws1, i, 4, row_data["Início Almoço"])
        style_data_cell(ws1, i, 5, row_data["Volta Almoço"])
        style_data_cell(ws1, i, 6, row_data["Saída"])
        style_data_cell(ws1, i, 7, row_data["Horas"])
        style_data_cell(ws1, i, 8, row_data["Projetos"])
        style_data_cell(ws1, i, 9, row_data["Empresas"])
        style_data_cell(ws1, i, 10, row_data["Observações"])

    # Total row
    total_row = len(preview_data) + 2
    ws1.cell(row=total_row, column=1, value="TOTAL").font = bold_font
    style_data_cell(ws1, total_row, 7, round(total_hours, 2))
    ws1.cell(row=total_row, column=7).font = bold_font

    # Column widths
    widths1 = [12, 6, 8, 12, 12, 8, 14, 30, 25, 30]
    for i, w in enumerate(widths1, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 2: Resumo por Projeto ──
    ws2 = wb.create_sheet("Resumo por Projeto")
    headers2 = ["Projeto", "Cor", "Horas Totais", "% do Total"]
    style_header(ws2, headers2)

    proj_summary = get_allocations_summary(start_date, end_date)
    for i, p in enumerate(proj_summary, 2):
        style_data_cell(ws2, i, 1, p["name"])
        style_data_cell(ws2, i, 2, p.get("color", ""))
        style_data_cell(ws2, i, 3, round(p["total_hours"], 2))
        pct = (p["total_hours"] / total_hours * 100) if total_hours > 0 else 0
        style_data_cell(ws2, i, 4, f"{pct:.1f}%")

    widths2 = [30, 10, 14, 12]
    for i, w in enumerate(widths2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 3: Resumo por Empresa ──
    ws_comp = wb.create_sheet("Resumo por Empresa")
    headers_comp = ["Empresa", "Cor", "Horas Totais", "% do Total"]
    style_header(ws_comp, headers_comp)

    comp_summary = get_company_allocations_summary(start_date, end_date)
    for i, c in enumerate(comp_summary, 2):
        style_data_cell(ws_comp, i, 1, c["name"])
        style_data_cell(ws_comp, i, 2, c.get("color", ""))
        style_data_cell(ws_comp, i, 3, round(c["total_hours"], 2))
        pct = (c["total_hours"] / total_hours * 100) if total_hours > 0 else 0
        style_data_cell(ws_comp, i, 4, f"{pct:.1f}%")

    for i, w in enumerate([30, 10, 14, 12], 1):
        ws_comp.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 4: Resumo Mensal ──
    ws3 = wb.create_sheet("Resumo Mensal")
    headers3 = ["Mês/Ano", "Dias Úteis", "Horas Trabalhadas", "Horas Esperadas", "Saldo"]
    style_header(ws3, headers3)

    # Group by month
    monthly = {}
    for e in enriched:
        d = date.fromisoformat(e["date"]) if isinstance(e["date"], str) else e["date"]
        key = d.strftime("%Y-%m")
        if key not in monthly:
            monthly[key] = {"hours": 0.0, "days": set()}
        monthly[key]["hours"] += e["worked_hours"]
        monthly[key]["days"].add(d.isoformat())

    from utils import count_business_days, get_setting
    daily = float(get_setting("daily_hours", "8"))

    for i, (key, data) in enumerate(sorted(monthly.items()), 2):
        y, m = key.split("-")
        label = date(int(y), int(m), 1).strftime("%b/%Y")
        bus_days = count_business_days(int(y), int(m))
        expected = bus_days * daily
        balance = data["hours"] - expected

        style_data_cell(ws3, i, 1, label)
        style_data_cell(ws3, i, 2, bus_days)
        style_data_cell(ws3, i, 3, round(data["hours"], 2))
        style_data_cell(ws3, i, 4, round(expected, 2))
        bal_cell = style_data_cell(ws3, i, 5, round(balance, 2))
        if balance < 0:
            bal_cell.font = Font(color="FF0000")
        else:
            bal_cell.font = Font(color="008000")

    widths3 = [14, 12, 16, 16, 10]
    for i, w in enumerate(widths3, 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    # Save
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ── Download Button ──
excel_data = generate_excel()
file_name = f"ponto_eletronico_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"

st.download_button(
    label="📥 Baixar Excel",
    data=excel_data,
    file_name=file_name,
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    type="primary",
    use_container_width=True,
)

st.caption(f"📁 Arquivo: `{file_name}`")
st.caption("📋 O Excel contém 4 abas: Registros Detalhados, Resumo por Projeto, Resumo por Empresa, e Resumo Mensal.")
